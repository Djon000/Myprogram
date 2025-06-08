import random
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Передача параметрів для алгоритму
# value_count обмеження кількості цінностей для предмету
def run_genetic_algorithm1(items, V, W, population_size, generations, mutation_rate, value_count):
    start_time = time.time()
    n = len(items)
    MAX_QUANTITY = value_count
    print("Запуск генетичного алгоритму")


    #chromosome - Список індексів предметів у списку items
    # items - список зі списками. У якому кожен підсписок представляє предмет у форматі
    # Назва, об'єм, цінність, вага
    def fitness(chromosome):
        # Задали початкову нульову загальну цінність, вагу та об'єм предметів
        total_value = 0
        total_volume = 0
        total_weight = 0
        # item_counts - словник, який підраховує скільки разів кожен тип предмета,
        # було вже взято
        item_counts = {}



        # Робимо перебір генів, проходимо по кожному гену(предмету) у хромосомі
        for item_index in chromosome:
            # Перевіряємо скільки разів вже було додано цей предмет,
            # якщо вже було досягнуто ліміту MAX_QUANTITY - то пропускаємо цей ген
            count = item_counts.get(item_index, 0)
            if count >= MAX_QUANTITY:
                continue
            # Беремо цінності предмета, яка залежить від порядку, наприклад items[i][2],
            # якщо такої цінності немає (предмет має менше цінностей ніж, count) то пропускаємо його.
            try:
                value = items[item_index][2 + count]
            except IndexError:
                continue
            # Додаємо параметри предметів до сумарної цінності, об'єму та ваги
            total_value += value
            total_volume += items[item_index][1]
            total_weight += items[item_index][-1]
            # Оновлюємо лічільник предмета, тим самим зберігаємо що 1 предмет такого типу
            # вже був використаний.
            item_counts[item_index] = count + 1
        # Уразі якщо перевищений об'єм або вага, хромосома є непридатної, і повертається return 0.
        if total_volume > V or total_weight > W:
            return 0
        # Якщо обмеження не порушені, то повертається сумарна цінність хромосом
        return total_value

    # Генеруємо випадкову хромосому, тобто рішення розміщення усіх предметів
    # n -кількість різних типів предметів, MAX_QUANTITY - їх цінності
    # n * MAX_QUANTITY - тут ми маємо деяку кількість генів у хромосомі
    # якщо n це 100, а MAX_QUANTITY - це 10 то 100*10 маємо випадкову довжину заповнення хромосоми
    # від 1 до 1000 генів у нашій хромосомі, тобто щось буде 0, щось 1, щось 10 предметів.
    def create_chromosome():
        return [random.randint(0, n - 1) for _ in range(random.randint(1, n * MAX_QUANTITY))]

    # Робимо кроссовер(схрещування) двох батьківських хромосом
    def crossover(parent1, parent2):
        # Знаходимо мінімальну довжину обох хромосом. (можуть мати різну довжину)
        min_length = min(len(parent1), len(parent2))
        # Якщо один із батьків має менше двох генів то не схрещуємо
        # Оскільки кроссовер потребує точки розрізу гену, то має бути мінімум 2 гени,
        # щоб між ними зробити розріз та схрещування, хромосома з 1 геном не ділиться.
        if min_length < 2:
            return parent1[:], parent2[:]
        # Випадкова точка розрізу в хромосомах батьків
        point = random.randint(1, min_length - 1)
        # Повертаємо дві нові хромосоми, які є результатом одноточкового схрещування
        # Двох батьківських хромосом. :point - означає що берем всі елементи з початку
        # списку до точки point.
        return parent1[:point] + parent2[point:], parent2[:point] + parent1[point:]



    #Виконуємо мутацію для хромосом
    def mutate(chromosome):
        # Перебираємо всі гени в хромосомі
        for i in range(len(chromosome)):
            # для кожного гену генеруємо випадкове число від 0 до 1, і якщо воно меньше
            # за mutation_rate(який було задано користувачем) то в такому разі
            # це значення змінюється на випадкове, у межах random.randint(0, n - 1)
            if random.random() < mutation_rate:
                chromosome[i] = random.randint(0, n - 1)
        # Знову генеруємо випадкове число, якщо воно меньше mutation_rate, то додаємо
        # новий елемент до хромосоми(випадковий у межах random.randint(0, n - 1))
        if random.random() < mutation_rate and len(chromosome) < n * MAX_QUANTITY:
            chromosome.append(random.randint(0, n - 1))
        # Знову генеруємо випадкове число, якщо воно меньше mutation_rate,
        # з хромосоми видаляється випадковий елемент
        if random.random() < mutation_rate and len(chromosome) > 1:
            chromosome.pop(random.randint(0, len(chromosome) - 1))
        return chromosome

    # Створення початкової популяції. Створили список population,
    # який містить population_size хромосом(індивідуальних).
    # Кожна хромосома генерується через виклик create_chromosome, яка створює
    # одну випадкову хромосому (розв'язання задачі)
    population = [create_chromosome() for _ in range(population_size)]
    # Ініціалізація глобального найкращого рішення.
    # З початку best_chromosome_ever - це пошук найкращого рішення у стартовій популяції
    # key=fitness - Вказує що ми порівнюємо за значеннями що повертає функція fitness.
    # Тобто ми отримуємо найкраще значення з усіх рішень для покоління
    best_chromosome_ever = max(population, key=fitness)
    # best_fitness_ever - Обчислює найкращу цінність з найбільшим значенням fitness
    best_fitness_ever = fitness(best_chromosome_ever)
    # best_generation - ініціалізує змінну, що зберігає номер покоління, в якому було знайдено
    # найкраще рішення.
    best_generation = 0

    # Виконує генерацію поколінь, деяку кількість разів, кожна генерація 1 нове покоління.
    for generation in range(generations):
        # Обчислимо придатність всіх хромосом
        # fitness_scores - Список що показує наскільки гарне кожне рішення
        fitness_scores = [fitness(chrom) for chrom in population]
        # total_fitness - сума всіх оцінок
        total_fitness = sum(fitness_scores)
        # Вибір батьків (селекція) Якщо if total_fitness == 0 Всі рішення дуже погані,
        # то вибираємо батьків випадково
        if total_fitness == 0:
            selected_population = random.choices(population, k=population_size)
        # Метод вибору за рулеткою score(береться з fitness_scores) / total_fitness - розраховуємо ймовірність
        # вибору кожної хромосоми пропорційно її fitness
        # на приклад хромосоми А = 2, B = 8, total_fitness = 10,
        # Значить хромосома B матиме шанс випадіння 0.8, а хромосома А лише 0.2
        # Якщо існують нормальні рішення, рахуємо ймовірність вибору кожної хромосоми
        # на основі її придатності(fitness), потім обираємо population_size хромосом
        # це і будуть батьки, наступного покоління.
        else:
            probabilities = [score / total_fitness for score in fitness_scores]
            selected_population = random.choices(population, weights=probabilities, k=population_size)



        #Наступне покоління
        next_generation = []
        # for i in range(0, population_size, 2): - Обходимо усю вибрану популяцію
        # з кроком 2, щоб можна було вибрати пари батьків
        for i in range(0, population_size, 2):
            #p1,p2 - батьки. min(i + 1, population_size - 1) - Умова, якщо population_size
            # є непарним, то останній батько буде спареним сам із собою.
            # Тобто забезпечуємо не вихід за межі списку.
            p1 = selected_population[i]
            p2 = selected_population[min(i + 1, population_size - 1)]
            # c1, c2 - Короссовер, тобто створення двох дітей, з пари батьків
            c1, c2 = crossover(p1, p2)
            # Мутації - обидва нащадки проходять мутацію, - випадкових змін в генах,
            # додавання, або видалення предметів. Результати мутацій додаються до нового покоління.
            next_generation.append(mutate(c1))
            next_generation.append(mutate(c2))
        # Завершення покоління, нове покоління, повністю замінює старе
        # :population_size - якщо було створено більше хромосом,
        # то обрізається до точної кількості
        population = next_generation[:population_size]



        # Оновлення глобального найкращого результату
        current_best = max(population, key=fitness)
        current_fitness = fitness(current_best)
        # Якщо поточне найраще рішення краще за попередні, тоді оновлюємо глобальну
        # найкращу хромосому best_chromosome_ever. Та зберігаємо її цінність best_fitness_ever.
        if current_fitness > best_fitness_ever:
            best_chromosome_ever = current_best
            best_fitness_ever = current_fitness
            # best_generation - запам'ятали покоління на якому було знайдено найкраще рішення
            best_generation = generation + 1

        progress = (generation + 1) / generations * 100
        print(f"Прогрес генетичного алгоритму: ({progress:.2f}%)")

    # Робота з найкращим хромосомом за всі покоління
    # Ці змінні потрібні для аналізу найкращого рішення після завершення всіх поколінь
    item_counts = {}
    total_value = 0
    total_volume = 0
    total_weight = 0
    print(" ")
    print("Предмети що були додані в трюм корабля:")
    print(f"{'Назва':<15}{'Шт.':<6}{'Цінність':<10}{'Вага':<8}{'Обʼєм':<8}")
    print("-" * 50)

    # Проходимо по кожному елементу (гену) в найкращій хромосомі(best_chromosome_ever)
    # Кожен елемент цієї хромосоми це індекс предмету зі списку items
    for item_index in best_chromosome_ever:
        # Отрмуємо кількість предметів певного типу, через item_index,
        # якщо предмет не було обрано встановлюємо кількість 0.
        count = item_counts.get(item_index, 0)
        # Перевірили чи не перевищує максимальну кількість предметів,
        # вже вибрана кількість предметів, якщо так, то пропускаємо предмет та
        # переходимо до наступного елемента.
        if count >= MAX_QUANTITY:
            continue
        # Для кожного предмета отримуємо його значення цінності
        try:
            value = items[item_index][2 + count]
        except IndexError:
            continue

        volume = items[item_index][1]
        weight = items[item_index][-1]
        name = items[item_index][0]

        print(f"{name:<15}{count + 1:<6}{value:<10}{weight:<8}{volume:<8}")

        total_value += value
        total_volume += volume
        total_weight += weight
        item_counts[item_index] = count + 1

    print("Максимальна цінність:", total_value)
    print("Загальний об'єм:", total_volume)
    print("Загальна вага:", total_weight)
    print("Найкращий результат знайдено на поколінні:", best_generation)

    execution_time = round(time.time() - start_time, 3)

    with open("result/genetic_algorithm_results.txt", "a", encoding="utf-8") as file:
        file.write(f"\nМаксимальна цінність: {total_value}\n")
        file.write(f"Загальний обʼєм: {total_volume}\n")
        file.write(f"Загальна вага: {total_weight}\n")
        file.write(f"Найкращий результат знайдено на поколінні: {best_generation}\n")
        file.write(f"Час виконання: {execution_time} секунд\n")
        file.write("-" * 50 + "\n\n")

        file.write("Предмети що були додані у трюм корабля:\n")
        file.write(f"{'Назва':<15}{'Шт.':<6}{'Цінність':<10}{'Обʼєм':<8}{'Вага':<8}\n")
        file.write("-" * 50 + "\n")
        item_counts = {}
        for item_index in best_chromosome_ever:
            count = item_counts.get(item_index, 0)
            if count >= MAX_QUANTITY:
                continue
            try:
                value = items[item_index][2 + count]
            except IndexError:
                continue
            volume = items[item_index][1]
            weight = items[item_index][-1]
            name = items[item_index][0]
            file.write(f"{name:<15}{count + 1:<6}{value:<10}{volume:<8}{weight:<8}\n")
            item_counts[item_index] = count + 1


    with open("result/genetic1_bestResults.txt", "a") as file:
        file.write(str(total_value) + ",")





    # Створення Excel-файлу з результатами
    wb = Workbook()
    ws = wb.active
    ws.title = "Результати GA"

    # Додаткові підсумкові значення
    ws.append([])  # Пустий рядок
    ws.append(["Загальні характеристики"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    cell = ws.cell(row=2, column=1)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.append(["Максимальна цінність", total_value])
    ws.append(["Загальний обʼєм", total_volume])
    ws.append(["Загальна вага", total_weight])
    ws.append(["Найкраще покоління", best_generation])
    ws.append(["Час виконання (с)", execution_time])
    ws.append([])  # Пустий рядок

    # Заголовки
    ws.append(["Назва", "Кількість", "Цінність", "Обʼєм", "Вага"])
    for col in range(1, 6):  # для стовпців з 1 по 5
        cell = ws.cell(row=9, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    item_counts = {}
    for item_index in best_chromosome_ever:
        count = item_counts.get(item_index, 0)
        if count >= MAX_QUANTITY:
            continue
        try:
            value = items[item_index][2 + count]
        except IndexError:
            continue
        volume = items[item_index][1]
        weight = items[item_index][-1]
        name = items[item_index][0]
        ws.append([name, count + 1, value, volume, weight])
        item_counts[item_index] = count + 1

    for col in range(1, 6):  # 1=A, 2=B, ..., 5=E
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20  # Встановлюємо ширину стовпця на 20

    # Збереження
    wb.save("result/genetic_algorithm_results.xlsx")

