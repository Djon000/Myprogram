import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# Отримує вхідні параметри
def run_greedy_algorithm(items, V, W, value_count):
    start_time = time.time() # Замірює час початку роботи
    print(" ")
    print("Запуск жадібного алгоритму")

    n = len(items)

    # Створення всіх можливих екземплярів предметів з відповідною цінністю
    expanded_items = []
    for idx, item in enumerate(items):# idx створюється ідентифікатор типу предмета для циклу
        #enumerate(items) Обійшли усі предмети зі списку items
        # Розпоковуємо список
        #  Витягнули name, volume, завдяки *values(Взяли всі елементи між 2 та перед останнім) ну і weight останній елемент
        name, volume, *values, weight = item
        # Створюємо не більше ніж value_count(Початкова задана кількість цінностей
        # що будуть зчитані) копій одного предмета
        # Але якщо в values(Фактично) меньше цінностей, ніж ми задали на початку, у value_count
        # то ми беремо лише ті цінності які є фактично, тобто values (якщо 2 цінності, то лише 2 а не 5)
        for i in range(min(value_count, len(values))):
            value = values[i]
            # Беремо цінність і, і якщо вона більше 0, то тоді беремо у розгорнутий список
            # це треба задля запобігання, додавання в список марних предметів
            if value > 0:
                # В розгорнутий список додаємо копію предмета, яка стає окремою одиницею у списку
                # тобто умовно окремою фізичною копією, то б то з 1 предмета у списку
                # з 10 цінностей, стає 10 предметів.
                expanded_items.append([name, volume, weight, value, idx, i])  # item_id + item_copy_number

    # Блок Перетворив модель з порядковими цінностями на простий список
    # що складається з окремих екземплярів.
    
    # Сортування за цінністю в порядку спадання
    #key=lambda x: x[3] - вказали критерії сортування елементу в списку
    #reverse = True - Задали сортування від меньшого до більшого(за спаданням)
    expanded_items.sort(key=lambda x: x[3], reverse=True)


    # Створили умовний трюм корабля
    backpack = []
    total_volume = 0
    total_weight = 0
    item_counters = {}  # Словник для відстеження
    # скільки вже взято предметів кожного типу

    processed_count = 0
    total_to_process = len(expanded_items)
    # Перебираємо всі елементи item у новому списку expanded_items,
    # де кожна цінність як окремий предмет
    for item in expanded_items:
        processed_count += 1
        progress_percent = (processed_count / total_to_process) * 100
        print(f"Прогрес жадібного алгоритму ({progress_percent:.1f}%)")        # Розпакували список у 6 змінних, взяли унікальний ID предмету(copy_index)
        name, volume, weight, value, item_id, copy_index = item
        # item_counters - Словник, який зберігає скільки копій кожного предмету вже покладено
        # .get(item_id, 0) - Якщо цього типу ще не було, значить 0 копій.
        # count_taken - тобто тут ми визначаємо які предмети покладені та скільки копій.
        count_taken = item_counters.get(item_id, 0)

        # Забезпечуємо послідовність вибору.Якщо ця копія не наступна за порядком — не беремо
        # Перевірка count_taken, якщо в count_taken, вже взято копій за індексом 0, а
        # ми зустрічаємо наприклад copy_index = 2, то ми цю копію не беремо. Тому що
        # нам підходить лише копія з індексом 0(яка є фактично першою).
        # Коли у нас буде співпадіння наприклад copy_index = 3 та count_taken = 3
        # Тоді все нормально, і умова не виконується, а копія береться.
        if copy_index != count_taken:
            # continue - повернулись в початок циклу(Перехід до наступного предмета)
            continue

        # Аналіз перевищення об'єму та ваги
        # total_volume + volume - загальний об'єм додати об'єм предмета <= Допустимому об'єму
        # total_weight + weight - загальна вага додати вагу предмета <= Допустимому вагу
        if total_volume + volume <= V and total_weight + weight <= W:
            # Якщо умову виконано, то додаємо, а значення загального об'єму та ваги оновлюємо
            backpack.append(item)
            total_volume += volume
            total_weight += weight
            # Оновлюємо лічільник копій для предмета що взяли
            item_counters[item_id] = count_taken + 1

    print(" ")
    print("Предмети що були додані в трюм корабля:")
    print(f"{'Назва':<15}{'Шт.':<6}{'Цінність':<10}{'Вага':<8}{'Обʼєм':<8}")
    print("-" * 50)
    # total_value Змінна для загальної цінності
    #  copy_index + 1  - показуємо, яка саме копія предмету 1ша, 2-га і так далі.
    total_value = 0
    for name, volume, weight, value, item_id, copy_index in backpack:
        print(f"{name:<15}{copy_index + 1:<6}{value:<10}{weight:<8}{volume:<8}")
        # total_value Підраховуємо загальну цінність
        total_value += value

    print("Загальна цінність речей в рюкзаку:", total_value)
    print("Загальний обʼєм:", total_volume)
    print("Загальна вага:", total_weight)
    # Виміряли час виконання алгоритму
    execution_time = round(time.time() - start_time, 3)

    # Запис результатів у файл
    with open("result/greedy_algorithm_results.txt", "w", encoding="utf-8") as file:
        file.write(f"\nМаксимальна цінність: {total_value}\n")
        file.write(f"Загальний обʼєм: {total_volume}\n")
        file.write(f"Загальна вага: {total_weight}\n")
        file.write(f"Час виконання алгоритму: {execution_time} сек.\n")
        file.write("-" * 50 + "\n\n")
        file.write("Предмети що були додані у трюм корабля:\n")
        file.write(f"{'Назва':<15}{'Шт.':<6}{'Цінність':<10}{'Обʼєм':<8}{'Вага':<8}\n")
        file.write("-" * 50 + "\n")
        for name, volume, weight, value, item_id, copy_index in backpack:
            file.write(f"{name:<15}{copy_index + 1:<6}{value:<10}{volume:<8}{weight:<8}\n")

    # Запис результатів у файл
    with open("result/greedy_times.txt", "a") as file:
        file.write(str(execution_time) + ",")

    # Запис результатів у файл
    with open("result/greedy_bestResults.txt", "a") as file:
        file.write(str(total_value) + ",")

    from openpyxl import Workbook

    # Створення Excel-файлу з результатами
    wb = Workbook()
    ws = wb.active
    ws.title = "Результати Greedy"

    ws.append([])
    ws.append(["Загальні характеристики"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.append(["Максимальна цінність", total_value])
    ws.append(["Загальний обʼєм", total_volume])
    ws.append(["Загальна вага", total_weight])
    ws.append(["Час виконання (с)", execution_time])
    ws.append([])

    # Заголовки
    ws.append(["Назва", "Кількість", "Цінність", "Обʼєм", "Вага"])

    item_counts = {}
    for name, volume, weight, value, item_id, copy_index in backpack:
        count = item_counts.get(item_id, 0)
        ws.append([name, copy_index + 1, value, volume, weight])
        item_counts[item_id] = count + 1

    for col in range(1, 6):  # 1=A, 2=B, ..., 5=E
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20  # Встановлюємо ширину стовпця на 20
    # Збереження
    wb.save("result/greedy_algorithm_results.xlsx")
    return backpack
